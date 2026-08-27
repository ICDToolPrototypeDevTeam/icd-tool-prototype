# -*- coding: utf-8 -*-
"""Forward completeness Excel detail report (Stage C8).

Three sheets:
  - 分析结果:      one row per business object (main fields, no long ID text)
  - 缺失HLR明细:  one row per (business object, missing HLR id)
  - 匹配证据明细: one row per (business object, matched HLR id)

The `_STATUS_LABELS` / `_final_status_label` / `_natural_reason` /
`_coverage_detail_row` helpers are shared with the Word generator.
"""

from __future__ import annotations

from pathlib import Path

from app.v4.models import ForwardBlocksOutput, ForwardCoverageOutput, ForwardCoverageResult

_STATUS_LABELS = {
    "covered_direct": "已覆盖",
    "covered_aggregate": "已覆盖（聚合）",
    "parent_referenced": "仅父级引用",
    "possible": "待确认",
    "uncovered": "未覆盖（疑似漏写）",
}


def _final_status_label(result: ForwardCoverageResult) -> str:
    """Unified display label: unsupported / input_error use analysis_status, the
    rest use coverage_status.
    """
    if result.analysis_status == "unsupported":
        return "不支持"
    if result.analysis_status == "input_error":
        return "输入异常"
    return _STATUS_LABELS.get(result.coverage_status, result.coverage_status or "—")


def _in_scene_clause(result: ForwardCoverageResult) -> str:
    """自然语言说明「在场候选为什么不能证明覆盖该对象」。

    不出现 weak_signal / trace_only / not_same_object 等内部术语，也不针对具体
    对象写死文案。
    """
    ai_verdict = result.ai_review.review_verdict if result.ai_review else ""
    rule_level = result.rule_level

    if ai_verdict == "not_same_object" or rule_level == "not_same_object":
        return "在场候选均非该对象"
    if rule_level == "parent_referenced":
        return "在场候选仅引用父级接口，未描述具体信号"
    if rule_level == "trace_only":
        return "在场候选与该对象无文本对应"
    if rule_level == "generic_signal":
        return "在场候选仅匹配到通用词"
    if rule_level == "weak_signal":
        return "在场候选仅匹配到单一短小片段"
    if ai_verdict == "unconfirmed":
        return "AI 复核无法确定在场候选是否描述该对象"
    return "在场候选无法证明覆盖该对象"


def _natural_reason(result: ForwardCoverageResult, missing_count: int) -> str:
    """用户可读的自然语言原因，缺失候选只显示数量，不展开完整 ID。"""
    as_status = result.analysis_status
    cs = result.coverage_status

    if as_status == "unsupported":
        return "该对象的协议类型（原生 A664）暂不支持分析。"
    if as_status == "input_error":
        return f"追溯表引用的候选 HLR 全部缺失于上传的 HLR 文档，无法分析（缺失 {missing_count} 条）。"

    if cs == "uncovered":
        if missing_count:
            return f"HLR 正文中未找到该对象的对应描述，疑似漏写；另有 {missing_count} 条追溯候选未上传。"
        return "HLR 正文中未找到该对象的对应描述，疑似漏写。"
    if cs == "parent_referenced":
        return "HLR 仅引用父级接口（端口/消息/Label），未描述具体信号，需人工确认。"
    if cs == "possible":
        clause = _in_scene_clause(result)
        if missing_count:
            return f"{clause}，另有 {missing_count} 条追溯候选未上传，暂无法判定是否漏写。"
        return f"{clause}，暂无法判定是否漏写，需人工确认。"
    if cs in ("covered_direct", "covered_aggregate"):
        return "已在 HLR 中找到该对象的对应描述。"
    return ""


def _coverage_detail_row(block, result: ForwardCoverageResult) -> dict:
    """Flatten a block + result into one report row (分析结果 sheet / Word list).

    只保留报告所需字段：对象身份（对象ID / 协议 / Label / 信号族 / 信号 / 设备）、
    最终状态（覆盖状态 / 原因）、缺失 HLR 数量。完整缺失 / 匹配 HLR ID 由 Excel
    的明细 Sheet 单独承载（Word 只显示数量）。
    """
    ident = block.identity
    missing = list(block.trace.missing_hlr_ids) if block.trace else []

    return {
        "business_object_id": result.business_object_id,
        "protocol": ident.protocol,
        "label": f"L{ident.label}" if ident.label else "",
        "signal_family": ident.signal_family,
        "signal": ident.signal,
        "device": ident.device or (", ".join(block.devices) if block.devices else ""),
        "missing_count": len(missing),
        "analysis_status": result.analysis_status,
        "coverage_status": result.coverage_status,
        "coverage_label": _final_status_label(result),
        "reason": _natural_reason(result, len(missing)),
    }


def _write_sheet(ws, headers: list[str], rows: list[list], widths: list[int]) -> None:
    """Apply header style, data rows, column widths, freeze first row, auto-filter."""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(name="微软雅黑", size=10, bold=True)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="微软雅黑", size=9)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    for r_idx, row in enumerate(rows, 2):
        for c_idx, v in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = cell_font
            cell.alignment = wrap_align

    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[_column_letter(idx)].width = w

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{_column_letter(len(headers))}{len(rows) + 1}"


def generate_forward_excel(
    coverage: ForwardCoverageOutput,
    blocks: ForwardBlocksOutput,
    output_path: Path,
) -> None:
    """Generate the three-sheet forward coverage Excel report."""
    from openpyxl import Workbook

    block_map = {b.business_object_id: b for b in blocks.blocks}
    pairs = [
        (block_map[r.business_object_id], r)
        for r in coverage.results
        if r.business_object_id in block_map
    ]

    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "分析结果"
    ws_missing = wb.create_sheet("缺失HLR明细")
    ws_matched = wb.create_sheet("匹配证据明细")

    # ── Sheet 1: 分析结果 ──
    headers_main = [
        "序号", "对象ID", "协议", "Label", "信号族", "信号", "设备",
        "覆盖状态", "原因", "缺失HLR数量",
    ]
    rows_main = []
    for i, (block, result) in enumerate(pairs, 1):
        row = _coverage_detail_row(block, result)
        rows_main.append([
            i, row["business_object_id"], row["protocol"], row["label"],
            row["signal_family"], row["signal"], row["device"],
            row["coverage_label"], row["reason"], row["missing_count"],
        ])
    _write_sheet(ws_main, headers_main, rows_main, [6, 30, 9, 9, 24, 24, 14, 14, 48, 10])

    # ── Sheet 2: 缺失HLR明细 ──
    headers_missing = ["序号", "对象ID", "协议", "缺失HLR ID"]
    rows_missing = []
    seq = 1
    for block, result in pairs:
        trace = block.trace
        if not trace:
            continue
        for mid in trace.missing_hlr_ids:
            rows_missing.append([seq, result.business_object_id, block.identity.protocol, mid])
            seq += 1
    _write_sheet(ws_missing, headers_missing, rows_missing, [6, 30, 9, 28])

    # ── Sheet 3: 匹配证据明细 ──
    headers_matched = ["序号", "对象ID", "协议", "匹配HLR ID"]
    rows_matched = []
    seq = 1
    for block, result in pairs:
        for hid in result.matched_hlr_ids:
            rows_matched.append([seq, result.business_object_id, block.identity.protocol, hid])
            seq += 1
    _write_sheet(ws_matched, headers_matched, rows_matched, [6, 30, 9, 28])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  Forward Excel: {output_path}")


def _column_letter(idx: int) -> str:
    """Convert a 1-based column index to an Excel column letter (supports >26)."""
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters
