# -*- coding: utf-8 -*-
"""Excel generator: EoICD itemization."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_HEADER_FONT = Font(name="微软雅黑", size=10, bold=True)
_HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_CELL_FONT = Font(name="微软雅黑", size=9)
_WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def generate_eoicd_excel(
    eoicd_json_path: Path,
    output_path: Path,
) -> None:
    """Generate EoICD itemization Excel workbook.

    Columns: 序号 | IRD ID | 描述
    """
    import json
    data = json.loads(eoicd_json_path.read_text(encoding="utf-8"))
    requirements = data.get("requirements", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "EoICD条目化清单"

    # Header row
    headers = ["序号", "IRD ID", "描述"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    # Data rows
    for i, req in enumerate(requirements, 1):
        ws.cell(row=i + 1, column=1, value=i).font = _CELL_FONT
        ws.cell(row=i + 1, column=2, value=req.get("ird_id", "")).font = _CELL_FONT
        ws.cell(row=i + 1, column=3, value=req.get("description", "")).font = _CELL_FONT
        ws.cell(row=i + 1, column=3).alignment = _WRAP_ALIGNMENT

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 90

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{len(requirements) + 1}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  EoICD Excel: {output_path}")
