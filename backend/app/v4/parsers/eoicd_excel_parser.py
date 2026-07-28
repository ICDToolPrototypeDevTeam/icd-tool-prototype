# -*- coding: utf-8 -*-
"""EoICD Excel parser — converts PubSub Excel tables to itemized requirements.

Follows generation_skill_v4.md rules:
- Processes both Publisher and Subscriber blocks per row
- Extracts only leaf (DP/RP) attributes
- Signal name includes leaf name
- Frame structure signal filtering (Rule 9)
- Subscriber dp_ref from same-row Publisher DP (Rule 8)
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

from app.v4.config import (
    DP_ATTRIBUTES_FOR_RP,
    get_display_name,
    get_unit,
    is_excluded,
    is_frame_signal,
)
from app.v4.models import EoICDOutput, EoICDRequirement


@dataclass
class LayerBlock:
    """Represents one layer's column range within a sheet."""

    layer_type: str
    start_col: int
    end_col: int = 0
    attr_names: list[str] = field(default_factory=list)


# ——— Row 1 / Row 2 scanning utilities ———


def _detect_side_boundaries(
    row_cells: list[Any],
) -> tuple[int | None, int | None]:
    """Find the column indices (1-indexed) for Publisher and Subscriber markers in Row 1."""
    pub_col = None
    sub_col = None
    for i, val in enumerate(row_cells, start=1):
        if val is not None:
            s = str(val).strip()
            if s == "Publisher" and pub_col is None:
                pub_col = i
            elif s == "Subscriber" and sub_col is None:
                sub_col = i
    return pub_col, sub_col


def _detect_layers(
    row_cells: list[Any],
    col_start: int,
    col_end: int,
) -> list[LayerBlock]:
    """Scan Row 2 within [col_start, col_end) to find layer blocks."""
    layers: list[LayerBlock] = []
    current: LayerBlock | None = None

    for i in range(col_start, col_end):
        val = row_cells[i - 1]
        if val is not None:
            if current is not None:
                current.end_col = i
                layers.append(current)
            current = LayerBlock(layer_type=str(val).strip(), start_col=i)

    if current is not None:
        current.end_col = col_end
        layers.append(current)

    return layers


def _read_attr_names(row_cells: list[Any], layers: list[LayerBlock]) -> None:
    """Populate each LayerBlock.attr_names from Row 3."""
    for layer in layers:
        for c in range(layer.start_col, layer.end_col):
            val = row_cells[c - 1]
            if val is not None:
                layer.attr_names.append(str(val).strip())
            else:
                layer.attr_names.append("")


def _read_layer_values(
    row_cells: list[Any], layer: LayerBlock
) -> dict[str, Any]:
    """Read attribute values for one layer from a data row."""
    values: dict[str, Any] = {}
    for offset, attr_name in enumerate(layer.attr_names):
        if not attr_name:
            continue
        c = layer.start_col + offset
        if c > len(row_cells):
            break
        val = row_cells[c - 1]
        if val is not None:
            values[attr_name] = val
    return values


def _is_empty(val: Any) -> bool:
    """Check if a value is empty (None, empty string, or whitespace-only string)."""
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    return False


# ——— Bus type parsing from FullName ———


_BUS_TYPE_PREFIX_PATTERNS = [
    ("A429", re.compile(r"^A429", re.IGNORECASE)),
    ("A664", re.compile(r"^A664", re.IGNORECASE)),
    ("A825", re.compile(r"^A825", re.IGNORECASE)),
    ("CAN", re.compile(r"^CAN", re.IGNORECASE)),
    ("AFDX", re.compile(r"^AFDX", re.IGNORECASE)),
    ("Analog", re.compile(r"^Analog", re.IGNORECASE)),
    ("Discrete", re.compile(r"^Discrete", re.IGNORECASE)),
]

# Substring patterns for bus type extraction from Tag / FullName (no ^ anchor).
# Word boundary prevents "CAN" matching inside words like "SCAN".
_SUBSTRING_BUS_TYPE_PATTERNS = [
    ("A429", re.compile(r"\bA429", re.IGNORECASE)),
    ("A664", re.compile(r"\bA664", re.IGNORECASE)),
    ("A825", re.compile(r"\bA825", re.IGNORECASE)),
    ("CAN", re.compile(r"\bCAN\b", re.IGNORECASE)),
    ("AFDX", re.compile(r"\bAFDX", re.IGNORECASE)),
    ("Analog", re.compile(r"\bAnalog", re.IGNORECASE)),
    ("Discrete", re.compile(r"\bDiscrete", re.IGNORECASE)),
]


def _parse_bus_type_from_fullname(fullname: str) -> str | None:
    """Extract bus type from an RP FullName string. Returns None if unrecognized.

    Checks prefix first, then falls back to substring search.
    """
    if not fullname:
        return None
    s = str(fullname)
    for bus_type, pattern in _BUS_TYPE_PREFIX_PATTERNS:
        if pattern.search(s):
            return bus_type
    for bus_type, pattern in _SUBSTRING_BUS_TYPE_PATTERNS:
        if pattern.search(s):
            return bus_type
    return None


def _parse_bus_type_from_tag(tag: str) -> str | None:
    """Extract bus type from LogicalPort Tag (e.g. 'A429Port' -> 'A429')."""
    if not tag:
        return None
    for bus_type, pattern in _SUBSTRING_BUS_TYPE_PATTERNS:
        if pattern.search(str(tag)):
            return bus_type
    return None


# ——— Signal name building ———


def _build_signal_name(layer_names: list[str], include_leaf: bool = True) -> str:
    """Build hierarchical signal name from layer Names with adjacent dedup.

    By default includes the leaf layer name (per generation_skill_v4 Rule 1).
    """
    end = len(layer_names) if include_leaf else len(layer_names) - 1
    if end <= 0:
        return ""
    parts = []
    for name in layer_names[:end]:
        name = str(name).strip() if name else ""
        if name and (not parts or parts[-1] != name):
            parts.append(name)
    return ".".join(parts)


# ——— Per-side leaf extraction ———


def _extract_layer_requirements(
    layers: list[LayerBlock],
    all_layer_names: list[str],
    all_layer_values: list[dict[str, Any]],
    side: str,  # "DP" or "RP"
    sheet_name: str,
    bus_type: str,
    source_label: str,  # "Publisher Table" or "Subscriber Table"
    is_dp_ref: bool = False,
) -> list[EoICDRequirement]:
    """Generate requirements from ALL layer attributes (not just leaf).

    Signal name for each layer = path from Software to that layer.
    """
    results: list[EoICDRequirement] = []

    for idx, layer in enumerate(layers):
        layer_vals = all_layer_values[idx]
        layer_name = all_layer_names[idx]
        if not layer_name:
            continue

        signal_name = _build_signal_name(all_layer_names[:idx + 1], include_leaf=True)
        if not signal_name:
            continue

        for attr_name, attr_value in layer_vals.items():
            if is_excluded(attr_name):
                continue
            if _is_empty(attr_value):
                continue

            display_name = get_display_name(attr_name, from_dp=is_dp_ref)
            unit = get_unit(attr_name)
            unit_str = str(unit) if unit else ""

            template = (
                "{signal}接收的{attr_display}应为{value}{unit}"
                if is_dp_ref
                else "{signal}的{attr_display}应为{value}{unit}"
            )
            desc = template.format(
                signal=signal_name,
                attr_display=display_name,
                value=attr_value,
                unit=unit_str,
            )

            results.append(
                EoICDRequirement(
                    ird_id="",
                    side=side,
                    sheet_name=sheet_name,
                    bus_type=bus_type,
                    layer_type=layer.layer_type,
                    attribute_name=attr_name,
                    attribute_value=attr_value,
                    unit=unit,
                    description=desc,
                    source=source_label,
                    is_dp_ref=is_dp_ref,
                    signal_name=signal_name,
                )
            )

    return results


# ——— Main parser ———


class EoICDExcelParser:
    """Parse both Publisher and Subscriber Excel files into merged requirement output.

    Processes both side blocks within each row. Extracts only leaf (DP/RP) attributes.
    Applies frame structure signal filtering (Rule 9) on Subscriber rows.
    """

    def __init__(
        self,
        publisher_path: Path | None = None,
        subscriber_path: Path | None = None,
    ):
        self.publisher_path = publisher_path
        self.subscriber_path = subscriber_path

    def parse(self) -> EoICDOutput:
        all_requirements: list[EoICDRequirement] = []
        sheet_stats: dict[str, dict[str, int]] = {}
        total_generated = 0
        total_raw = 0

        # Build DP lookup from Publisher file (for cross-file dp_ref)
        pub_dp_lookup: dict[str, dict[str, Any]] = {}

        if self.publisher_path:
            pub_reqs, pub_gen, pub_raw, pub_stats, pub_dps = self._parse_one_file(
                self.publisher_path,
                primary_side="Publisher",
                leaf_type="DP",
                source_label="Publisher Table",
                pub_dp_lookup=pub_dp_lookup,
            )
            all_requirements.extend(pub_reqs)
            total_generated += pub_gen
            total_raw += pub_raw
            sheet_stats.update(pub_stats)
            pub_dp_lookup.update(pub_dps)

        if self.subscriber_path:
            sub_reqs, sub_gen, sub_raw, sub_stats, _ = self._parse_one_file(
                self.subscriber_path,
                primary_side="Subscriber",
                leaf_type="RP",
                source_label="Subscriber Table",
                pub_dp_lookup=pub_dp_lookup,
            )
            all_requirements.extend(sub_reqs)
            total_generated += sub_gen
            total_raw += sub_raw
            for k, v in sub_stats.items():
                if k in sheet_stats:
                    for layer, cnt in v.items():
                        sheet_stats[k][layer] = sheet_stats[k].get(layer, 0) + cnt
                else:
                    sheet_stats[k] = v

        # P2.3: DP-RP conflict resolution — DP wins for same (signal_name, attr_name)
        conflict_map: dict[tuple[str, str], list[EoICDRequirement]] = defaultdict(list)
        for req in all_requirements:
            conflict_map[(req.signal_name, req.attribute_name)].append(req)
        resolved: list[EoICDRequirement] = []
        for entries in conflict_map.values():
            dp_entries = [e for e in entries if e.side == "DP"]
            rp_entries = [e for e in entries if e.side == "RP"]
            if dp_entries and rp_entries:
                resolved.extend(dp_entries)
            else:
                resolved.extend(entries)
        all_requirements = resolved

        # Deduplicate
        dedup_set: set[tuple] = set()
        deduped: list[EoICDRequirement] = []
        for req in all_requirements:
            key = (
                req.side,
                req.layer_type,
                req.signal_name,
                req.attribute_name,
                str(req.attribute_value),
                req.is_dp_ref,
            )
            if key not in dedup_set:
                dedup_set.add(key)
                deduped.append(req)

        # Assign global entry IDs
        for idx, req in enumerate(deduped, start=1):
            layer_abbr = req.layer_type[:6]
            req.ird_id = f"IRD-{req.bus_type}-{layer_abbr}-{idx:04d}"

        dup_removed = total_generated - len(deduped)
        return EoICDOutput(
            total_generated=total_generated,
            total_raw=total_raw,
            total_after_dedup=len(deduped),
            duplicates_removed=dup_removed,
            sheet_statistics=sheet_stats,
            requirements=deduped,
        )

    def _parse_one_file(
        self,
        file_path: Path,
        primary_side: str,  # "Publisher" or "Subscriber"
        leaf_type: str,  # "DP" or "RP"
        source_label: str,
        pub_dp_lookup: dict[str, dict[str, Any]],
    ) -> tuple[
        list[EoICDRequirement], int, int, dict[str, dict[str, int]], dict[str, dict[str, Any]]
    ]:
        """Parse one Excel file.

        Returns: (requirements, total_generated, total_raw, sheet_stats, updated_pub_dp_lookup)
        """
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

        all_reqs: list[EoICDRequirement] = []
        sheet_stats: dict[str, dict[str, int]] = {}
        total_generated = 0
        total_raw = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Read all rows at once; read_only mode doesn't support max_row/max_column.
            rows: list[list] = [list(row) for row in ws.iter_rows(values_only=True)]
            if len(rows) < 4:
                continue

            max_col = max(len(r) for r in rows) if rows else 0
            total_cols = max_col + 1
            row1 = rows[0] if len(rows) > 0 else []
            row2 = rows[1] if len(rows) > 1 else []
            row3 = rows[2] if len(rows) > 2 else []

            pub_col, sub_col = _detect_side_boundaries(row1)

            # --- Publisher block ---
            pub_layers: list[LayerBlock] = []
            if pub_col:
                pub_end = sub_col if sub_col else total_cols
                pub_layers = _detect_layers(row2, pub_col, pub_end)
                _read_attr_names(row3, pub_layers)

            # --- Subscriber block ---
            sub_layers: list[LayerBlock] = []
            if sub_col:
                sub_layers = _detect_layers(row2, sub_col, total_cols)
                _read_attr_names(row3, sub_layers)

            bus_type = sheet_name.split("-")[0] if "-" in sheet_name else sheet_name
            sheet_raw = 0
            layer_counts: dict[str, int] = defaultdict(int)
            row_dedup: set[tuple] = set()

            for r in range(4, len(rows) + 1):
                row_cells = rows[r - 1]

                # --- Parse Publisher block ---
                pub_names: list[str] = []
                pub_values: list[dict[str, Any]] = []
                dp_attrs: dict[str, Any] = {}
                dp_name = ""

                if pub_layers:
                    for layer in pub_layers:
                        vals = _read_layer_values(row_cells, layer)
                        pub_values.append(vals)
                        name = vals.get("Name", "")
                        pub_names.append(str(name).strip() if name else "")
                        if layer.layer_type == "DP":
                            dp_attrs = vals
                            dp_name = pub_names[-1]

                # --- Parse Subscriber block ---
                sub_names: list[str] = []
                sub_values: list[dict[str, Any]] = []

                if sub_layers:
                    for layer in sub_layers:
                        vals = _read_layer_values(row_cells, layer)
                        sub_values.append(vals)
                        name = vals.get("Name", "")
                        sub_names.append(str(name).strip() if name else "")

                # --- Generate requirements ---

                # Publisher DP requirements
                if pub_layers and primary_side == "Publisher":
                    # Skip if leaf DP Name is empty
                    if dp_name:
                        reqs = _extract_layer_requirements(
                            pub_layers, pub_names, pub_values,
                            side="DP",
                            sheet_name=sheet_name, bus_type=bus_type,
                            source_label=source_label,
                        )
                        for req in reqs:
                            total_generated += 1
                            key = ("DP", req.layer_type, req.signal_name,
                                   req.attribute_name, str(req.attribute_value))
                            if key not in row_dedup:
                                row_dedup.add(key)
                                all_reqs.append(req)
                                sheet_raw += 1
                                layer_counts[req.layer_type] += 1

                    # Collect DP into lookup for cross-file use
                    dp_guid = dp_attrs.get("Guid", "")
                    if dp_guid:
                        pub_dp_lookup[str(dp_guid)] = dp_attrs

                # Subscriber RP requirements
                if sub_layers and primary_side == "Subscriber":
                    # Rule 9: frame structure filtering on DP Name
                    if dp_name and is_frame_signal(dp_name):
                        continue

                    # Skip if leaf RP Name is empty
                    rp_name = sub_names[-1] if sub_names else ""
                    if not rp_name:
                        continue

                    # P2.1: Parse RP actual bus type — Tag priority, FullName fallback
                    rp_bus_type = bus_type
                    # ① Try LogicalPort Tag first
                    lp_tag = ""
                    for layer, vals in zip(sub_layers, sub_values):
                        if layer.layer_type == "LogicalPort":
                            lp_tag = str(vals.get("Tag", "")).strip()
                            break
                    if lp_tag:
                        parsed = _parse_bus_type_from_tag(lp_tag)
                        if parsed:
                            rp_bus_type = parsed
                    # ② Fall back to RP FullName
                    if rp_bus_type == bus_type and sub_values:
                        rp_fullname = str(sub_values[-1].get("FullName", ""))
                        if rp_fullname:
                            parsed = _parse_bus_type_from_fullname(rp_fullname)
                            if parsed:
                                rp_bus_type = parsed

                    # RP own attributes (all layers)
                    reqs = _extract_layer_requirements(
                        sub_layers, sub_names, sub_values,
                        side="RP",
                        sheet_name=sheet_name, bus_type=rp_bus_type,
                        source_label=source_label,
                    )

                    # P2.4: DP→RP attribute merge — inject same-row Publisher
                    # block attributes into RP entries so profiles carry full
                    # ICD attributes instead of just Label + SDIExpected.
                    if pub_layers:
                        # Collect DP attributes per Publisher DP layer,
                        # preserving the pub_name so sub-signal identity is
                        # available downstream (sub_signals in signal_profiler).
                        dp_layer_attrs: list[tuple[str, dict[str, Any]]] = []
                        dp_attr_map: dict[str, Any] = {}
                        for idx, layer in enumerate(pub_layers):
                            if layer.layer_type != "DP":
                                continue
                            layer_name = pub_names[idx] if idx < len(pub_names) else ""
                            layer_vals: dict[str, Any] = {}
                            for attr_name, attr_value in pub_values[idx].items():
                                if is_excluded(attr_name):
                                    continue
                                if _is_empty(attr_value):
                                    continue
                                layer_vals[attr_name] = attr_value
                                dp_attr_map[attr_name] = attr_value
                            if layer_vals:
                                dp_layer_attrs.append((layer_name, layer_vals))

                        rp_signal_name = _build_signal_name(
                            sub_names, include_leaf=True
                        )

                        # When DP layer has no attributes for this row,
                        # RP non-Label/SDI values are unreliable — discard
                        # them to avoid polluting profiles with stale RP data.
                        if not dp_attr_map:
                            reqs = [r for r in reqs
                                    if r.attribute_name in ("Label", "SDIExpected")]

                        rp_has_attr = {req.attribute_name for req in reqs}

                        # (a) Override RP non-authoritative attributes with DP data.
                        # Only Label + SDIExpected use RP's own data source;
                        # all other attributes (BitOffsetWithinDS, ParameterSize,
                        # DataFormatType, etc.) use same-row DP data.
                        for req in reqs:
                            attr = req.attribute_name
                            if attr in ("Label", "SDIExpected"):
                                continue
                            dp_val = dp_attr_map.get(attr)
                            if dp_val is not None:
                                req.attribute_value = dp_val
                                unit = get_unit(attr)
                                unit_str = str(unit) if unit else ""
                                dn = get_display_name(attr, from_dp=False)
                                req.description = (
                                    f"{rp_signal_name}的{dn}"
                                    f"应为{dp_val}{unit_str}"
                                )

                        # (b) Union: DP attributes RP doesn't have, plus per-DP-layer
                        # DataFormatType/ParameterSize (needed for per-sub-signal
                        # size/dtype when one RP signal covers multiple DP sub-fields).
                        for layer_name, layer_vals in dp_layer_attrs:
                            for attr_name, attr_value in layer_vals.items():
                                if attr_name in rp_has_attr:
                                    if attr_name not in ("DataFormatType", "ParameterSize"):
                                        continue
                                if attr_name in ("Label", "SDIExpected"):
                                    continue
                                unit = get_unit(attr_name)
                                unit_str = str(unit) if unit else ""
                                dn = get_display_name(attr_name, from_dp=False)
                                desc = (
                                    f"{rp_signal_name}的{dn}"
                                    f"应为{attr_value}{unit_str}"
                                )
                                reqs.append(EoICDRequirement(
                                    ird_id="",
                                    side="RP",
                                    sheet_name=sheet_name,
                                    bus_type=rp_bus_type,
                                    layer_type="RP",
                                    attribute_name=attr_name,
                                    attribute_value=attr_value,
                                    unit=unit,
                                    description=desc,
                                    source=source_label,
                                    is_dp_ref=True,
                                    dp_ref_name=layer_name,
                                    signal_name=rp_signal_name,
                                ))

                    for req in reqs:
                        total_generated += 1
                        key = ("RP", req.layer_type, req.signal_name,
                               req.attribute_name, str(req.attribute_value),
                               req.dp_ref_name)
                        if key not in row_dedup:
                            row_dedup.add(key)
                            all_reqs.append(req)
                            sheet_raw += 1
                            layer_counts[req.layer_type] += 1

                    # P2.2: Extract full DP attributes from pub_layers, bound to DP signal name
                    # Use pub_names (DP hierarchy) so Label-based clustering works correctly.
                    # The existing dp_ref also uses pub_names for the same reason.
                    dp_signal_name = _build_signal_name(pub_names, include_leaf=True)
                    if pub_layers and dp_name and dp_signal_name:
                        for idx, layer in enumerate(pub_layers):
                            layer_vals = pub_values[idx]
                            layer_name = pub_names[idx]
                            if not layer_name:
                                continue
                            for attr_name, attr_value in layer_vals.items():
                                if is_excluded(attr_name):
                                    continue
                                if _is_empty(attr_value):
                                    continue

                                display_name = get_display_name(attr_name, from_dp=False)
                                unit = get_unit(attr_name)
                                unit_str = str(unit) if unit else ""
                                desc = (
                                    f"{dp_signal_name}的{display_name}"
                                    f"应为{attr_value}{unit_str}"
                                )

                                total_generated += 1
                                key = ("DP", layer.layer_type, dp_signal_name,
                                       attr_name, str(attr_value))
                                if key not in row_dedup:
                                    row_dedup.add(key)
                                    all_reqs.append(
                                        EoICDRequirement(
                                            ird_id="",
                                            side="DP",
                                            sheet_name=sheet_name,
                                            bus_type=rp_bus_type,
                                            layer_type=layer.layer_type,
                                            attribute_name=attr_name,
                                            attribute_value=attr_value,
                                            unit=unit,
                                            description=desc,
                                            source=source_label,
                                            is_dp_ref=False,
                                            signal_name=dp_signal_name,
                                        )
                                    )
                                    sheet_raw += 1
                                    layer_counts[layer.layer_type] += 1

                    # Rule 8: dp_ref — extract configured DP attributes (kept for backward compat)
                    if dp_attrs:
                        dp_ref_values = {
                            attr: dp_attrs[attr]
                            for attr in DP_ATTRIBUTES_FOR_RP
                            if attr in dp_attrs and not _is_empty(dp_attrs.get(attr))
                        }
                        if dp_ref_values:
                            # Use pub_names for signal (from Publisher side hierarchy)
                            signal_name = _build_signal_name(pub_names, include_leaf=True)
                            if signal_name:
                                for attr_name, attr_value in dp_ref_values.items():
                                    total_generated += 1
                                    display_name = get_display_name(attr_name, from_dp=True)
                                    unit = get_unit(attr_name)
                                    unit_str = str(unit) if unit else ""
                                    desc = (
                                        f"{signal_name}接收的{display_name}"
                                        f"应为{attr_value}{unit_str}"
                                    )
                                    key = ("DP", "DP", signal_name, attr_name, str(attr_value))
                                    if key not in row_dedup:
                                        row_dedup.add(key)
                                        all_reqs.append(
                                            EoICDRequirement(
                                                ird_id="",
                                                side="DP",
                                                sheet_name=sheet_name,
                                                bus_type=rp_bus_type,
                                                layer_type="DP",
                                                attribute_name=attr_name,
                                                attribute_value=attr_value,
                                                unit=unit,
                                                description=desc,
                                                source=source_label,
                                                is_dp_ref=True,
                                                signal_name=signal_name,
                                            )
                                        )
                                        sheet_raw += 1
                                        layer_counts["DP"] += 1

            total_raw += sheet_raw
            if layer_counts:
                sheet_stats[sheet_name] = dict(layer_counts)

        return all_reqs, total_generated, total_raw, sheet_stats, pub_dp_lookup
