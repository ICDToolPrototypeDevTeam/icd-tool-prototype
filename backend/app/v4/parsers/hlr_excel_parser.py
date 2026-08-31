# -*- coding: utf-8 -*-
"""HLR Excel parser — extracts software high-level requirements from .xlsx sheets.

Adapted from the RPDU local branch to fit the V4 profile-driven architecture.
Outputs the same ``HLROutput`` shape as ``HLRWordParser`` so downstream
stages (labeling, matching, judging) stay parser-agnostic.

Expected Excel layout (Sheet1):
  Row 1: title row (merged cells, skipped)
  Row 2: column headers (需求编号 | 模块名称 | 需求内容 | ...)
  Row 3+: data rows

Column mapping (1-based):
  A (col 0) -> requirement_id          需求编号
  B (col 1) -> implementation_method    模块名称
  C (col 2) -> content                  需求内容
  Remaining fields default to empty strings (Excel format has no
  direct equivalent columns).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from app.v4.models import HLROutput, HLRRequirement
from app.v4.parsers.hlr_parser_base import HLRParserBase

# Data starts at row 3 (skip Row 1 title + Row 2 header).
_DATA_START_ROW = 3

# Column index map (0-based)
_COL_REQUIREMENT_ID = 0      # A: 需求编号
_COL_MODULE_NAME = 1          # B: 模块名称 -> implementation_method
_COL_CONTENT = 2              # C: 需求内容


def _cell_str(value) -> str:
    """Safely extract cell text; None/empty uniformly yield empty string."""
    if value is None:
        return ""
    return str(value).strip()


class HLRExcelParser(HLRParserBase):
    """Parse HLR rows from an Excel sheet into a standard ``HLROutput``."""

    def parse(self) -> HLROutput:
        wb = openpyxl.load_workbook(str(self.source_path), data_only=True)
        ws = wb[wb.sheetnames[0]]

        requirements: list[HLRRequirement] = []

        for row in ws.iter_rows(
            min_row=_DATA_START_ROW, max_row=ws.max_row, values_only=True
        ):
            requirement_id = (
                _cell_str(row[_COL_REQUIREMENT_ID])
                if len(row) > _COL_REQUIREMENT_ID else ""
            )
            content = (
                _cell_str(row[_COL_CONTENT])
                if len(row) > _COL_CONTENT else ""
            )
            module_name = (
                _cell_str(row[_COL_MODULE_NAME])
                if len(row) > _COL_MODULE_NAME else ""
            )

            # Skip empty rows (no requirement_id or no content).
            if not requirement_id or not content:
                continue

            requirements.append(
                HLRRequirement(
                    requirement_id=requirement_id,
                    content=content,
                    object_type="",
                    is_derived="",
                    rationale="",
                    is_safety_related="",
                    verification_method="",
                    implementation_method=module_name,
                    source_file=self.source_name,
                )
            )

        wb.close()

        return HLROutput(
            source_file=self.source_name,
            total_count=len(requirements),
            requirements=requirements,
            glossary=[],  # Excel format has no glossary table.
        )