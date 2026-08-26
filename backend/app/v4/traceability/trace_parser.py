# -*- coding: utf-8 -*-
"""Traceability parser: reads traceability Excel tables and builds
an HLR -> ICD BlockKey index for reverse-matching pre-filtering.

Profile-driven: file patterns, sheet selection, and column positions are
all controlled by TraceabilityConfig.

Two parsing strategies are supported:

  - ``profile_columns`` (legacy, default): read columns by their hardcoded
    numeric index from ``TraceabilityTableConfig.columns``. Used by
    AMS/FGMC/HSCU whose traceability sheets have a stable layout.

  - ``header_adaptive`` (RPDU, opt-in via ``ControllerProfile.trace_strategy``):
    scan each sheet's header row for keywords ("ERD编号", "ICD FullName",
    "需求编号", "HLR", "下层需求", ...) and pick the sheet with the most
    matches. Column positions are not configured. Useful when traceability
    sheets vary by revision and the column order is unreliable.
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from app.v4.profiles.base import (
    ControllerProfile,
    TraceabilityConfig,
    TraceabilityTableConfig,
)


_LABEL_SEGMENT_RE = re.compile(r"^(L\d+)", re.IGNORECASE)
_SIGNAL_FAMILY_PREFIX_RE = re.compile(r"^L\d+_(?:\d+_)?[A-Z]+\d+_")
_PROTOCOL_BLOCKKEY_SUFFIXES = ("/SDI", "/LABEL", "/PARITY", "/SSM", "/OCTLBL")

# Keyword lists used by the adaptive header scanner (RPDU).
_TABLE1_ERD_KEYWORDS = ("erd编号", "erd", "er需求编号", "关联的erd", "需求编号")
_TABLE1_ICD_KEYWORDS = ("icd fullname", "icd_fullname", "icd full name",
                        "icd名称", "接口名称", "fullname")
_TABLE2_ERD_KEYWORDS = ("erd编号", "erd", "需求编号", "er需求编号")
_TABLE2_HLR_KEYWORDS = ("hlr", "下级需求", "高层需求", "软件高层需求")
_TABLE2_MODULE_KEYWORDS = ("模块名称", "模块", "module")
_EMPTY_CELL_SENTINELS = ("none", "#n/a", "")


@dataclass
class TraceabilityIndex:
    """Pre-computed index mapping HLR IDs to ICD BlockKeys."""

    hlr_to_blocks: dict[str, set[str]] = field(default_factory=dict)
    hlr_to_erds: dict[str, list[str]] = field(default_factory=dict)
    erd_to_icd: dict[str, list[str]] = field(default_factory=dict)

    total_hlrs_traced: int = 0
    total_erds: int = 0
    total_icd_fullnames: int = 0
    icd_mapped_to_blocks: int = 0
    icd_unmapped: list[str] = field(default_factory=list)


def name_to_block_key(name: str) -> str | None:
    """Convert an ICD FullName or signal_name to its ICDBlock.block_key.

    Mirrors signal_profiler._extract_profile_key + _extract_signal_family.
    """
    if not name or not name.strip():
        return None
    segments = [s.strip() for s in name.split(".") if s.strip()]
    if not segments:
        return None
    label = None
    for seg in segments:
        m = _LABEL_SEGMENT_RE.match(seg)
        if m:
            label = m.group(1)
            break
    leaf = segments[-1]
    if label:
        label_value = label[1:]
        m2 = _SIGNAL_FAMILY_PREFIX_RE.match(leaf)
        if m2:
            family = leaf[m2.end():]
        else:
            family = leaf
        return f"L{label_value}/{family}"
    else:
        if len(segments) >= 2:
            parent = segments[-2]
            return f"{parent}/{leaf}"
        return leaf


def _select_sheet(wb, cfg: TraceabilityTableConfig):
    """Select sheet by name keyword match, then fallback to index."""
    if cfg.sheet_match.by_name_keywords:
        for kw in cfg.sheet_match.by_name_keywords:
            for sname in wb.sheetnames:
                if kw in sname:
                    return wb[sname]
    if cfg.sheet_match.fallback_index < len(wb.sheetnames):
        return wb[wb.sheetnames[cfg.sheet_match.fallback_index]]
    return wb[wb.sheetnames[0]]


def _read_table2_erd_to_hlr(
    fpath: Path, cfg: TraceabilityTableConfig
) -> dict[str, list[str]]:
    """Read Table 2 (parent<->HLR matrix) to build parent -> HLR mapping.

    Returns: {parent_id: [hlr_id, ...]}
    """
    import openpyxl

    erd_to_hlr: dict[str, list[str]] = defaultdict(list)
    if not fpath.exists():
        raise FileNotFoundError(f"Table 2 file not found: {fpath}")

    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = _select_sheet(wb, cfg)

    col_erd = cfg.columns["erd"]
    col_hlr = cfg.columns["hlr"]
    col_module = cfg.columns.get("module", -1)

    # Header / data split is now controlled by cfg.data_start_row, so the
    # iteration starts at (data_start_row + 1) -- row index is 1-based in openpyxl.
    data_start = cfg.data_start_row

    current_parent = ""
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start + 1, max_row=ws.max_row, values_only=True),
        start=data_start + 1,
    ):
        parent_cell = str(row[col_erd]).strip() if len(row) > col_erd and row[col_erd] else ""
        hlr_cell = str(row[col_hlr]).strip() if len(row) > col_hlr and row[col_hlr] else ""
        module_name = ""
        if col_module >= 0 and len(row) > col_module and row[col_module]:
            module_name = str(row[col_module]).strip()

        # Fill-forward: non-empty parent updates current
        if parent_cell and parent_cell != "None":
            current_parent = parent_cell

        if not current_parent:
            continue

        # Skip rows whose module matches skip_module list (case-insensitive)
        if module_name and module_name.upper() in {m.upper() for m in cfg.skip_module}:
            continue

        if not hlr_cell or hlr_cell == "None":
            continue

        erd_to_hlr[current_parent].append(hlr_cell)

    wb.close()
    return dict(erd_to_hlr)


def _read_table1_erd_to_icd(
    fpath: Path, cfg: TraceabilityTableConfig
) -> dict[str, list[str]]:
    """Read Table 1 (ERD<->ICD mapping) to build parent -> ICD FullName mapping."""
    import openpyxl

    erd_to_icd: dict[str, list[str]] = defaultdict(list)
    if not fpath.exists():
        raise FileNotFoundError(f"Table 1 file not found: {fpath}")

    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = _select_sheet(wb, cfg)

    col_erd = cfg.columns["erd"]
    col_icd = cfg.columns["icd_fullname"]

    current_erd = ""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        erd_cell = str(row[col_erd]).strip() if len(row) > col_erd and row[col_erd] else ""
        icd_fn = str(row[col_icd]).strip() if len(row) > col_icd and row[col_icd] else ""

        if erd_cell and erd_cell != "None":
            current_erd = erd_cell

        if not current_erd:
            continue

        if not icd_fn or icd_fn == "None":
            continue

        erd_to_icd[current_erd].append(icd_fn)

    wb.close()
    return dict(erd_to_icd)


def _locate_header_row(ws) -> tuple[list[str], int] | None:
    """Scan the first 5 rows of a sheet for a real header row.

    WPS exports often have 1-2 title / empty rows before the actual
    header. The first row containing any non-empty cell is taken as the
    candidate header. Returns ``(headers_lower, header_row_index)`` or
    ``None`` if the sheet is empty.
    """
    for check_row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if check_row and any(c for c in check_row if c not in (None, "")):
            headers_lower = [
                str(h).strip().lower() if h else "" for h in check_row
            ]
            return headers_lower, check_row[0].row if hasattr(check_row[0], "row") else 0
    return None


def _find_col_by_keywords(
    headers: list[str], keywords: tuple[str, ...]
) -> int | None:
    """Return the 0-based column index whose header text contains any of
    the given keywords (case-insensitive substring match). ``None`` when
    no header matches.
    """
    for i, h in enumerate(headers):
        if not h:
            continue
        for kw in keywords:
            if kw in h:
                return i
    return None


def _cell_is_blank(value) -> bool:
    """True if a cell is missing, None, '#N/A', or 'none'."""
    if value is None:
        return True
    s = str(value).strip()
    return s.lower() in _EMPTY_CELL_SENTINELS


def _read_table1_header_adaptive(fpath: Path) -> dict[str, list[str]]:
    """Read Table 1 with header-keyword adaptive column detection (RPDU).

    Strategy:
      1. Scan every sheet for the row containing ERD and ICD-FullName
         column headers (substring match against ``_TABLE1_ERD_KEYWORDS`` /
         ``_TABLE1_ICD_KEYWORDS``).
      2. Use the sheet with the most ERD→ICD mappings.
      3. Read rows from that sheet, fill-forward ERD across blank cells
         (merged-cell layout).

    Supports original 设备需求与系统ICD追溯表 layouts, the EoICD baseline
    table (接口基线表_EoICD), and any future format whose header row
    carries ERD / ICD FullName labels.
    """
    import openpyxl

    if not fpath.exists():
        raise FileNotFoundError(f"Table 1 file not found: {fpath}")

    wb = openpyxl.load_workbook(fpath, data_only=True)
    best_result: dict[str, list[str]] = {}
    best_sheet = ""

    for name in wb.sheetnames:
        ws = wb[name]
        located = _locate_header_row(ws)
        if located is None:
            continue
        headers_lower, _hdr_row = located

        erd_col = _find_col_by_keywords(headers_lower, _TABLE1_ERD_KEYWORDS)
        icd_col = _find_col_by_keywords(headers_lower, _TABLE1_ICD_KEYWORDS)
        if erd_col is None or icd_col is None:
            continue

        sheet_result: dict[str, list[str]] = defaultdict(list)
        current_erd = ""
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            erd_val = (
                str(row[erd_col]).strip()
                if len(row) > erd_col and not _cell_is_blank(row[erd_col])
                else ""
            )
            icd_val = (
                str(row[icd_col]).strip()
                if len(row) > icd_col and not _cell_is_blank(row[icd_col])
                else ""
            )
            if erd_val:
                current_erd = erd_val
            if not current_erd or not icd_val:
                continue
            sheet_result[current_erd].append(icd_val)

        if len(sheet_result) > len(best_result):
            best_result = dict(sheet_result)
            best_sheet = name

    wb.close()
    if best_result:
        print(
            f"  [trace] Table 1: sheet='{best_sheet}', "
            f"{len(best_result)} ERDs mapped (header-adaptive)"
        )
    return best_result


def _read_table2_header_adaptive(
    fpath: Path, cfg: TraceabilityTableConfig | None = None,
) -> dict[str, list[str]]:
    """Read Table 2 with header-keyword adaptive column detection (RPDU).

    Same strategy as ``_read_table1_header_adaptive`` but for the
    ERD↔HLR matrix. When ``cfg`` is provided, rows whose module column
    matches ``cfg.skip_module`` are skipped (case-insensitive), so the
    header-adaptive path still honours per-profile exclusions such as
    ``EICD`` for AMS.
    """
    import openpyxl

    if not fpath.exists():
        raise FileNotFoundError(f"Table 2 file not found: {fpath}")

    skip_module_upper = (
        {m.upper() for m in cfg.skip_module} if cfg is not None else set()
    )

    wb = openpyxl.load_workbook(fpath, data_only=True)
    best_result: dict[str, list[str]] = {}
    best_sheet = ""

    for name in wb.sheetnames:
        ws = wb[name]
        located = _locate_header_row(ws)
        if located is None:
            continue
        headers_lower, _hdr_row = located

        erd_col = _find_col_by_keywords(headers_lower, _TABLE2_ERD_KEYWORDS)
        hlr_col = _find_col_by_keywords(headers_lower, _TABLE2_HLR_KEYWORDS)
        if erd_col is None or hlr_col is None:
            continue
        mod_col = _find_col_by_keywords(headers_lower, _TABLE2_MODULE_KEYWORDS)

        sheet_result: dict[str, list[str]] = defaultdict(list)
        current_erd = ""
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            erd_val = (
                str(row[erd_col]).strip()
                if len(row) > erd_col and not _cell_is_blank(row[erd_col])
                else ""
            )
            hlr_val = (
                str(row[hlr_col]).strip()
                if len(row) > hlr_col and not _cell_is_blank(row[hlr_col])
                else ""
            )
            mod_val = ""
            if mod_col is not None and len(row) > mod_col and not _cell_is_blank(row[mod_col]):
                mod_val = str(row[mod_col]).strip()

            if erd_val:
                current_erd = erd_val
            if not current_erd or not hlr_val:
                continue
            if hlr_val == current_erd:
                # Skip self-referential rows (e.g. parent ERD = self).
                continue
            if skip_module_upper and mod_val and mod_val.upper() in skip_module_upper:
                continue
            sheet_result[current_erd].append(hlr_val)

        if len(sheet_result) > len(best_result):
            best_result = dict(sheet_result)
            best_sheet = name

    # Fallback: hardcoded RPDU column layout (col A=ERD, col D=HLR, col E=module,
    # data from row 4). Used when both header columns share the same text
    # (e.g. "需求编号") and the section header in row 1 ("当前需求文档"/"下层需求文档")
    # does not match _TABLE2_*_KEYWORDS. Matches the column layout that RPDU's
    # 适配RPDU需求格式代码 originally hardcoded for the
    # ``单模块需求矩阵分析(设备2软件)`` workbook.
    if not best_result:
        for name in wb.sheetnames:
            ws = wb[name]
            current_erd = ""
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
                erd_cell = str(row[0]).strip() if row[0] else ""
                hlr_cell = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                module_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                if erd_cell and erd_cell != "None":
                    current_erd = erd_cell
                if not current_erd:
                    continue
                if skip_module_upper and module_name and module_name.upper() in skip_module_upper:
                    continue
                if not hlr_cell or hlr_cell == "None":
                    continue
                if hlr_cell == current_erd:
                    continue
                best_result.setdefault(current_erd, []).append(hlr_cell)
            if best_result:
                best_sheet = name
                break

    wb.close()
    if best_result:
        print(
            f"  [trace] Table 2: sheet='{best_sheet}', "
            f"{len(best_result)} ERDs mapped (header-adaptive)"
        )
    return best_result


def _discover_trace_files(
    trace_dir: Path, cfg: TraceabilityConfig,
) -> tuple[Path, Path, Path | None]:
    """Locate the two traceability Excel files via glob patterns (RPDU).

    Returns (table1, table2, bridge_or_None). The bridge file is the
    optional 3rd Excel used by ``build_trace_index`` when the ERD ID
    namespaces of Table 1 and Table 2 don't overlap (e.g. RPDUDRD_* vs
    EPDSSRS_*) — discovered by filename heuristics inherited from the
    original RPDU ``适配RPDU需求格式代码`` package.
    """
    pool = sorted(p for p in trace_dir.glob("*.xlsx") if p.is_file())

    def _find_one(patterns: tuple[str, ...]) -> Path | None:
        for pat in patterns:
            for p in pool:
                if p.match(pat):
                    return p
        return None

    t1 = _find_one(cfg.table1.filename_patterns)
    t2 = _find_one(cfg.table2.filename_patterns)

    if t1 is None:
        raise FileNotFoundError(
            f"Cannot locate Table 1 (patterns {cfg.table1.filename_patterns}) in {trace_dir}"
        )
    if t2 is None:
        raise FileNotFoundError(
            f"Cannot locate Table 2 (patterns {cfg.table2.filename_patterns}) in {trace_dir}"
        )

    # Optional 3rd file: bridge table for cross-namespace ERD mapping
    # (RPDUDRD_* ↔ EPDSSRS_*). Discovered by filename keyword after Table 1
    # and Table 2 are pinned; falls back to "any remaining file".
    bridge = None
    for p in pool:
        if p == t1 or p == t2:
            continue
        if any(kw in p.name for kw in ["系统2设备", "系统", "桥接", "bridge"]):
            bridge = p
            break
    if bridge is None:
        remaining = [p for p in pool if p != t1 and p != t2]
        if remaining:
            bridge = remaining[0]

    return t1, t2, bridge


def _read_bridge_table(
    fpath: Path, t1_erds: set[str], t2_erds: set[str],
) -> dict[str, list[str]]:
    """Read bridge table to map Table 2's ERD IDs to Table 1's ERD IDs (RPDU).

    Bridge table links two requirement ID namespaces (e.g. EPDSSRS ↔ RPDUDRD).
    Uses prefix matching to determine which column belongs to which table.

    Returns ``{t2_er_id: [t1_er_id, ...]}``. Inherited from the original
    RPDU ``适配RPDU需求格式代码`` package — needed when Table 1 ERD IDs
    (system-level, e.g. ``EPDSSRS_*``) and Table 2 ERD IDs (device-level,
    e.g. ``RPDUDRD_*``) live in different namespaces.
    """
    import openpyxl

    if not fpath.exists():
        return {}

    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    def _get_prefixes(erd_set: set[str]) -> set[str]:
        prefixes: set[str] = set()
        for erd in erd_set:
            if "_" in erd:
                prefixes.add(erd.split("_")[0].upper())
        return prefixes

    t1_prefixes = _get_prefixes(t1_erds)
    t2_prefixes = _get_prefixes(t2_erds)
    print(
        f"  [trace] Bridge: T1 prefixes={t1_prefixes}, T2 prefixes={t2_prefixes}",
    )

    # Scan rows to find columns with ID values
    col_samples: dict[int, set[str]] = defaultdict(set)
    for row in ws.iter_rows(min_row=3, max_row=min(100, ws.max_row), values_only=True):
        for i, cell in enumerate(row[:8]):
            val = str(cell).strip() if cell else ""
            if val and val not in ("None", "#N/A", ""):
                col_samples[i].add(val)

    # Find which column matches T1 prefixes and which matches T2 prefixes
    t1_col = None
    t2_col = None
    best_t1_match = 0
    best_t2_match = 0

    for col, vals in col_samples.items():
        if len(vals) < 2:
            continue
        col_prefixes: set[str] = set()
        for v in vals:
            if "_" in v:
                col_prefixes.add(v.split("_")[0].upper())
        t1_match = len(col_prefixes & t1_prefixes)
        t2_match = len(col_prefixes & t2_prefixes)
        if t1_match > best_t1_match:
            best_t1_match = t1_match
            t1_col = col
        if t2_match > best_t2_match:
            best_t2_match = t2_match
            t2_col = col

    if t1_col is None or t2_col is None or t1_col == t2_col:
        wb.close()
        print(
            f"  [trace] Bridge: no column match (t1_col={t1_col}, t2_col={t2_col})",
        )
        return {}

    print(
        f"  [trace] Bridge: T1 col={t1_col} (prefix match={best_t1_match}), "
        f"T2 col={t2_col} (prefix match={best_t2_match})",
    )

    t2_to_t1: dict[str, list[str]] = defaultdict(list)
    current_t1 = ""
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        t1_val = str(row[t1_col]).strip() if len(row) > t1_col and row[t1_col] else ""
        t2_val = str(row[t2_col]).strip() if len(row) > t2_col and row[t2_col] else ""

        if t1_val and t1_val not in ("None", "#N/A", ""):
            current_t1 = t1_val
        if not current_t1:
            continue
        if not t2_val or t2_val in ("None", "#N/A", ""):
            continue

        t2_to_t1[t2_val].append(current_t1)

    wb.close()
    print(f"  [trace] Bridge: mapped {len(t2_to_t1)} T2 ERDs → T1 ERDs")
    return dict(t2_to_t1)


def build_trace_index(
    trace_dir: Path,
    cfg: TraceabilityConfig,
    profile: ControllerProfile | None = None,
) -> TraceabilityIndex:
    """Build the complete HLR -> BlockKey traceability index.

    Dispatches between two parsing strategies:

      - ``"profile_columns"`` (default) — read columns by their hardcoded
        numeric index in ``TraceabilityTableConfig.columns``. AMS/FGMC/HSCU.
      - ``"header_adaptive"`` — locate ERD/HLR/ICD FullName columns by
        scanning header rows for Chinese keywords. RPDU.

    Pass ``profile=None`` to force the default ``profile_columns`` path
    (byte-identical to the pre-#74 / pre-#63 behaviour).
    """
    strategy = (profile.trace_strategy if profile is not None else "profile_columns")
    t1_path, t2_path, bridge_path = _discover_trace_files(trace_dir, cfg)

    if strategy == "header_adaptive":
        erd_to_hlr = _read_table2_header_adaptive(t2_path, cfg.table2)
        erd_to_icd = _read_table1_header_adaptive(t1_path)
    else:
        erd_to_hlr = _read_table2_erd_to_hlr(t2_path, cfg.table2)
        erd_to_icd = _read_table1_erd_to_icd(t1_path, cfg.table1)

    # Optional 3-table bridge (RPDU): when T1 and T2 ERD namespaces differ
    # (e.g. EPDSSRS_* vs RPDUDRD_*), the bridge table maps T2 → T1 so the
    # HLR→ERD→ICD chain can complete.
    t2_erds = set(erd_to_hlr.keys())
    t1_erds = set(erd_to_icd.keys())
    bridge_map: dict[str, list[str]] = {}
    if not (t2_erds & t1_erds) and bridge_path and bridge_path.exists():
        print(
            f"  [trace] No ERD overlap between Table 1 and Table 2; "
            f"using bridge table: {bridge_path.name}"
        )
        bridge_map = _read_bridge_table(bridge_path, t1_erds, t2_erds)

    index = TraceabilityIndex()
    index.hlr_to_erds = defaultdict(list)
    index.erd_to_icd = erd_to_icd
    index.total_erds = len(erd_to_hlr)

    hlr_to_icd: dict[str, set[str]] = defaultdict(set)
    for erd, hlr_list in erd_to_hlr.items():
        icd_list = list(erd_to_icd.get(erd, []))
        if bridge_map and not icd_list:
            for upper_erd in bridge_map.get(erd, []):
                icd_list.extend(erd_to_icd.get(upper_erd, []))
        for hlr in hlr_list:
            index.hlr_to_erds[hlr].append(erd)
            for icd_fn in icd_list:
                hlr_to_icd[hlr].add(icd_fn)

    index.total_hlrs_traced = len(hlr_to_icd)

    all_icd_fullnames: set[str] = set()
    for icd_set in hlr_to_icd.values():
        all_icd_fullnames.update(icd_set)
    index.total_icd_fullnames = len(all_icd_fullnames)

    index.hlr_to_blocks = defaultdict(set)
    protocol_skipped = 0
    for icd_fn in all_icd_fullnames:
        bk = name_to_block_key(icd_fn)
        if not bk:
            index.icd_unmapped.append(icd_fn)
            continue
        if bk.endswith(_PROTOCOL_BLOCKKEY_SUFFIXES):
            protocol_skipped += 1
            continue
        index.icd_mapped_to_blocks += 1
        for hlr, icd_set in hlr_to_icd.items():
            if icd_fn in icd_set:
                index.hlr_to_blocks[hlr].add(bk)
    if protocol_skipped:
        print(f"  Trace index: skipped {protocol_skipped} protocol-overhead block(s)")

    index.hlr_to_blocks = dict(index.hlr_to_blocks)
    index.hlr_to_erds = dict(index.hlr_to_erds)
    return index