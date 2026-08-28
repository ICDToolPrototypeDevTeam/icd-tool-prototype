# -*- coding: utf-8 -*-
"""Forward traceability scope builder (Stage C2).

Determines the analysis scope for forward completeness checking:

  - "trace" mode: scope = ICD FullNames from Table1 (设备→ICD), each linked to
    candidate HLR IDs from Table2 (设备→高层需求) via the shared ERD编号.
    An ERD present in Table1 but absent from Table2 yields a scope item with
    empty candidate_hlr_ids (=> this EoICD object has no traced HLR reference).
  - "full" mode: scope = ALL leaf EoICD business objects; candidate HLRs are
    resolved later by inverted-index recall (no cartesian product). No trace
    files are read in this mode.

Validation (clear errors, never a raw IndexError/FileNotFoundError):
  - trace mode requires both device_icd_trace_file (Table1) and
    system_device_trace_file (Table2).
  - sheet presence + column count are validated before reading.
  - trace-referenced HLR IDs missing from the parsed HLR set are surfaced as
    input_errors (non-fatal: analysis continues with matched HLRs only).
"""

from __future__ import annotations

from pathlib import Path

from app.v4.models import (
    EoICDOutput,
    ForwardScopeItem,
    ForwardScopeOutput,
    HLROutput,
)
from app.v4.traceability.trace_parser import (
    _read_table1_erd_to_icd,
    _read_table2_erd_to_hlr,
    name_to_block_key,
)

# Trace table structure (fixed; no config):
#   Table1 (device_icd_trace_file)  -> sheet index 1, col D=ERD, col H=ICD FullName
#   Table2 (system_device_trace_file) -> sheet index 0, col A=ERD, col D=HLR, col E=module
_TABLE1_SHEET_INDEX = 1
_TABLE1_MIN_COLUMNS = 8      # need col H (index 7)
_TABLE2_SHEET_INDEX = 0
_TABLE2_MIN_COLUMNS = 5      # need col E (index 4)


def _validate_trace_table(
    fpath: Path,
    field: str,
    sheet_index: int,
    min_columns: int,
    errors: list[dict],
) -> bool:
    """Validate that a trace file exists and has the expected sheet/columns."""
    if fpath is None or not fpath.exists():
        errors.append({
            "kind": "trace_file_missing",
            "field": field,
            "detail": f"{field} is required in trace mode but was not provided or not found: {fpath}",
        })
        return False
    try:
        import openpyxl

        wb = openpyxl.load_workbook(fpath, data_only=True)
        sheets = wb.sheetnames
        if len(sheets) <= sheet_index:
            errors.append({
                "kind": "trace_sheet_missing",
                "field": field,
                "detail": f"{field} has {len(sheets)} sheet(s); expected sheet index {sheet_index}",
            })
            wb.close()
            return False
        ws = wb[sheets[sheet_index]]
        if ws.max_column < min_columns:
            errors.append({
                "kind": "trace_columns_missing",
                "field": field,
                "detail": f"{field} sheet '{sheets[sheet_index]}' has {ws.max_column} column(s); expected >= {min_columns}",
            })
            wb.close()
            return False
        wb.close()
        return True
    except Exception as exc:  # noqa: BLE001 — surface any openpyxl error clearly
        errors.append({
            "kind": "trace_open_error",
            "field": field,
            "detail": f"failed to open {field}: {type(exc).__name__}: {exc}",
        })
        return False


def _read_trace_table(
    fpath: Path | None,
    field: str,
    reader,
    sheet_index: int,
    min_columns: int,
    errors: list[dict],
) -> dict[str, list[str]]:
    """Validate then read one trace table into {erd_id: [value, ...]}."""
    if not _validate_trace_table(fpath, field, sheet_index, min_columns, errors):
        return {}
    try:
        return reader(Path(fpath))
    except (IndexError, KeyError, FileNotFoundError) as exc:
        errors.append({
            "kind": "trace_parse_error",
            "field": field,
            "detail": f"failed to read {field}: {exc}",
        })
        return {}
    except Exception as exc:  # noqa: BLE001
        errors.append({
            "kind": "trace_parse_error",
            "field": field,
            "detail": f"failed to read {field}: {type(exc).__name__}: {exc}",
        })
        return {}


def _locate_eoicd_signals(
    eoicd: EoICDOutput,
    scope_fullnames: list[str],
) -> dict[str, list[str]]:
    """Map each scope FullName to the parsed EoICD signal_names it locates.

    Uses name_to_block_key (the shared, decoupled FullName/signal_name
    normalizer) so the mapping is a best-effort diagnostic, not the forward
    stable identity (that is derived later by block_builder).
    """
    # block_key -> [signal_name] over leaf entries only (DP/RP).
    key_to_signals: dict[str, list[str]] = {}
    for req in eoicd.requirements:
        if req.layer_type not in ("DP", "RP"):
            continue
        key = name_to_block_key(req.signal_name)
        if not key:
            continue
        key_to_signals.setdefault(key, []).append(req.signal_name)

    result: dict[str, list[str]] = {}
    for fn in scope_fullnames:
        key = name_to_block_key(fn)
        located = key_to_signals.get(key, []) if key else []
        result[fn] = sorted(set(located))
    return result


def build_forward_scope(
    eoicd: EoICDOutput,
    hlr: HLROutput,
    analysis_mode: str = "full",
    device_icd_trace_file: Path | None = None,
    system_device_trace_file: Path | None = None,
) -> ForwardScopeOutput:
    """Build the forward analysis scope.

    Args:
        eoicd: parsed EoICD output (Stage C1).
        hlr: parsed HLR output (Stage C1).
        analysis_mode: "trace" or "full".
        device_icd_trace_file: Table1 (设备→ICD), required in trace mode.
        system_device_trace_file: Table2 (设备→高层需求), required in trace mode.
    """
    errors: list[dict] = []

    if analysis_mode == "full":
        return ForwardScopeOutput(
            analysis_mode="full",
            scope_source="all leaf EoICD business objects (DP/RP)",
            total_scope_fullnames=0,
            total_candidate_hlrs=0,
            scope_items=[],
            input_errors=errors,
        )

    # —— trace mode ——
    erd_to_icd = _read_trace_table(
        device_icd_trace_file,
        "device_icd_trace_file",
        _read_table1_erd_to_icd,
        _TABLE1_SHEET_INDEX,
        _TABLE1_MIN_COLUMNS,
        errors,
    )
    erd_to_hlr = _read_trace_table(
        system_device_trace_file,
        "system_device_trace_file",
        _read_table2_erd_to_hlr,
        _TABLE2_SHEET_INDEX,
        _TABLE2_MIN_COLUMNS,
        errors,
    )

    # Parsed HLR id set for missing-reference detection.
    parsed_hlr_ids = {r.requirement_id for r in hlr.requirements}

    scope_items: list[ForwardScopeItem] = []
    seen_fullnames: dict[str, ForwardScopeItem] = {}
    unmatched_hlr_ids: set[str] = set()

    for erd, fullnames in erd_to_icd.items():
        candidate_hlr_ids = list(erd_to_hlr.get(erd, []))
        for hlr_id in candidate_hlr_ids:
            if hlr_id and hlr_id not in parsed_hlr_ids:
                unmatched_hlr_ids.add(hlr_id)
        for fn in fullnames:
            fn = (fn or "").strip()
            if not fn:
                continue
            if fn not in seen_fullnames:
                seen_fullnames[fn] = ForwardScopeItem(
                    icd_fullname=fn,
                    erd_ids=[],
                    candidate_hlr_ids=[],
                )
            item = seen_fullnames[fn]
            if erd not in item.erd_ids:
                item.erd_ids.append(erd)
            # Merge candidate HLRs (dedup, keep order).
            merged = list(dict.fromkeys(item.candidate_hlr_ids + candidate_hlr_ids))
            item.candidate_hlr_ids = merged

    scope_items = list(seen_fullnames.values())

    # 正向缺陷修正 #4：逐 item 记录 raw 候选里缺失于上传 HLR 文档的 id。
    # raw candidate_hlr_ids 保留用于审计；仅 analyzable（候选 - 缺失）参与匹配/AI。
    for item in scope_items:
        item.missing_hlr_ids = [
            h for h in item.candidate_hlr_ids if h not in parsed_hlr_ids
        ]

    if unmatched_hlr_ids:
        errors.append({
            "kind": "missing_hlr_input",
            "detail": (
                f"{len(unmatched_hlr_ids)} HLR id(s) referenced by the trace tables "
                f"were not found in the uploaded HLR document; first 10: "
                f"{sorted(unmatched_hlr_ids)[:10]}"
            ),
        })

    located = _locate_eoicd_signals(eoicd, [s.icd_fullname for s in scope_items])
    missing_fullnames: list[str] = []
    for item in scope_items:
        item.located_eoicd_signal_names = located.get(item.icd_fullname, [])
        if not item.located_eoicd_signal_names:
            missing_fullnames.append(item.icd_fullname)

    if missing_fullnames:
        errors.append({
            "kind": "missing_icd_fullname",
            "detail": (
                f"{len(missing_fullnames)} ICD FullName(s) from the trace tables "
                f"could not be located in the parsed EoICD; first 10: "
                f"{missing_fullnames[:10]}"
            ),
        })

    total_candidate_hlrs = len({h for s in scope_items for h in s.candidate_hlr_ids})

    return ForwardScopeOutput(
        analysis_mode="trace",
        scope_source="device→ICD trace table (Table1) linked to device→HLR trace table (Table2) via ERD编号",
        total_scope_fullnames=len(scope_items),
        total_candidate_hlrs=total_candidate_hlrs,
        scope_items=scope_items,
        input_errors=errors,
    )
