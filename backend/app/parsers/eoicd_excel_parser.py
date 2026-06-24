"""
EoICD Excel 附件解析器。

解析 Publisher Table / Subscriber Table 格式的 Excel 文件，
提取各总线类型 Sheet 的结构化数据（层级链、Publisher 侧、Subscriber 侧）。
"""

from pathlib import Path

import openpyxl

from app.models import EoICDExcelSheet, ParsedEoICDExcel


def _detect_bus_type(sheet_name: str) -> str:
    """从 Sheet 名中检测总线类型。"""
    name_upper = sheet_name.upper()
    for bus in ("A664", "A825", "A429", "ANALOG", "DISCRETE"):
        if bus in name_upper:
            if bus == "ANALOG":
                return "Analog"
            if bus == "DISCRETE":
                return "Discrete"
            return bus
    return ""


def _find_pub_sub_split(ws) -> int:
    """通过 Row 1 找到 Publisher 和 Subscriber 的分割列 (1-indexed)。

    返回 Subscriber 起始列。若无法检测返回 -1。
    """
    row1_values = []
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=col).value or "").strip().lower()
        row1_values.append(val)

    # 查找 "subscriber" 首次出现的位置
    for i, v in enumerate(row1_values):
        if "subscriber" in v:
            return i + 1  # 1-indexed

    # 备选：通过合并单元格定位
    for merge in ws.merged_cells.ranges:
        if merge.min_row == 1 and merge.min_col > 1:
            val = str(ws.cell(row=1, column=merge.min_col).value or "").strip().lower()
            if "subscriber" in val:
                return merge.min_col

    return -1


def _parse_entity_groups(ws, pub_sub_split: int) -> list[dict]:
    """解析 Row 2-3 的实体层级链和列范围。

    通过 Row 2 的合并单元格确定每个实体的列范围，
    通过 Row 3 获取每个实体下的具体字段名。

    Returns:
        [
            {"entity": "Software", "col_start": 1, "col_end": 22, "fields": ["Tag","Name",...]},
            {"entity": "LogicalPort", "col_start": 23, "col_end": 37, "fields": [...]},
            ...
        ]
    """
    # 收集 Row 2 的合并单元格范围作为实体定义
    entity_merges = []
    for merge in ws.merged_cells.ranges:
        if merge.min_row == 2 and merge.max_row == 2:
            entity_merges.append({
                "col_start": merge.min_col,
                "col_end": merge.max_col,
                "name": str(ws.cell(row=2, column=merge.min_col).value or "").strip(),
            })

    # 按列排序
    entity_merges.sort(key=lambda m: m["col_start"])

    # 为每个实体收集 Row 3 的字段名
    entities = []
    for em in entity_merges:
        fields = []
        for col in range(em["col_start"], em["col_end"] + 1):
            val = str(ws.cell(row=3, column=col).value or "").strip()
            if val and val != "None":
                fields.append(val)
        entities.append({
            "entity": em["name"],
            "col_start": em["col_start"],
            "col_end": em["col_end"],
            "fields": fields,
        })

    return entities


def _parse_data_rows(ws, start_row: int, pub_sub_split: int, entities: list[dict]) -> list[dict]:
    """解析数据行（从 start_row 开始）。

    每行按实体组拆分，每列值以 "entity.field" 为 key 存入 dict。
    跳过所有列为空的行。
    """
    rows = []
    for row_idx in range(start_row, ws.max_row + 1):
        row_data = {}
        has_value = False

        for ent in entities:
            entity_name = ent["entity"]
            for i, field in enumerate(ent["fields"]):
                col = ent["col_start"] + i
                if col > ws.max_column:
                    break
                val = ws.cell(row=row_idx, column=col).value
                if val is not None:
                    str_val = str(val).strip()
                    if str_val and str_val.lower() not in ("none", ""):
                        has_value = True
                        row_data[f"{entity_name}.{field}"] = str_val

        if has_value:
            rows.append(row_data)

    return rows


def _parse_sheet(ws, sheet_name: str) -> EoICDExcelSheet:
    """解析单个 Sheet 为 EoICDExcelSheet。"""
    bus_type = _detect_bus_type(sheet_name)
    pub_sub_split = _find_pub_sub_split(ws)

    if pub_sub_split < 0:
        # 无法确定分割点：整表作为一个 side 处理
        raise ValueError(f"Sheet '{sheet_name}': 无法检测 Publisher/Subscriber 分割点")

    # 分离 Publisher 侧和 Subscriber 侧的实体
    all_entities = _parse_entity_groups(ws, pub_sub_split)
    pub_entities = [e for e in all_entities if e["col_start"] < pub_sub_split]
    sub_entities = [e for e in all_entities if e["col_start"] >= pub_sub_split]

    # 提取层级链（仅保留实体名，去重）
    pub_chain = [e["entity"] for e in pub_entities if e["entity"]]
    sub_chain = [e["entity"] for e in sub_entities if e["entity"]]

    # Publisher 侧 headers（Row 3 字段名，仅 Publisher 侧列）
    pub_headers = []
    for ent in pub_entities:
        pub_headers.extend(ent["fields"])

    # Subscriber 侧 headers
    sub_headers = []
    for ent in sub_entities:
        sub_headers.extend(ent["fields"])

    # 查找数据起始行（跳过前 3 行表头）
    data_start = 4
    for r in range(4, min(ws.max_row + 1, 10)):
        has_data = False
        for col in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(row=r, column=col).value
            if val is not None and str(val).strip():
                has_data = True
                break
        if has_data:
            data_start = r
            break

    # 解析数据行
    publisher_rows = _parse_data_rows(ws, data_start, pub_sub_split, pub_entities)
    subscriber_rows = _parse_data_rows(ws, data_start, pub_sub_split, sub_entities)

    return EoICDExcelSheet(
        sheet_name=sheet_name,
        bus_type=bus_type,
        publisher_headers=pub_headers,
        subscriber_headers=sub_headers,
        publisher_rows=publisher_rows,
        subscriber_rows=subscriber_rows,
        hierarchy_chain=pub_chain,
    )


def parse_eoicd_excel(excel_paths: list[Path]) -> ParsedEoICDExcel:
    """解析一个或多个 EoICD Excel 附件。

    支持 Publisher Table 和 Subscriber Table 两种格式的 Excel 文件。
    每个文件包含 5 种总线类型的 Sheet（A664-RP, A825-RP, A429-RP, Analog-RP, Discrete-RP）。

    Args:
        excel_paths: EoICD Excel 文件路径列表

    Returns:
        ParsedEoICDExcel（包含所有文件的 Sheet 解析结果）
    """
    all_sheets: list[EoICDExcelSheet] = []
    source_files: list[str] = []

    for xlsx_path in excel_paths:
        if not xlsx_path.exists():
            continue

        source_files.append(str(xlsx_path))
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 4:
                continue

            try:
                sheet = _parse_sheet(ws, sheet_name)
                if sheet.publisher_rows or sheet.subscriber_rows:
                    all_sheets.append(sheet)
            except ValueError as e:
                # 无法解析的 Sheet 跳过后继续
                print(f"WARNING: {e}")

        wb.close()

    return ParsedEoICDExcel(
        source_files=source_files,
        sheets=all_sheets,
    )
