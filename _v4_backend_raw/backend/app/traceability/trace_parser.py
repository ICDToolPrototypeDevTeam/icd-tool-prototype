# -*- coding: utf-8 -*-
"""Traceability parser: reads two Excel traceability tables and builds
an HLR -> ICD BlockKey index for reverse-matching pre-filtering.

Trace chains:
  Table 2 (设备->高层): ERD编号 -> HLR ID
  Table 1 (ICD->设备):   ERD编号 -> ICD FullName
  Combined:             HLR ID -> ICD FullName -> BlockKey
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path


# -- Regex patterns (mirror signal_profiler.py logic) -----------------

_LABEL_SEGMENT_RE = re.compile(r'^(L\d+)', re.IGNORECASE)
_SIGNAL_FAMILY_PREFIX_RE = re.compile(r'^L\d+_(?:\d+_)?[A-Z]+\d+_')


# -- Data structures -------------------------------------------------

@dataclass
class TraceabilityIndex:
    """Pre-computed index mapping HLR IDs to ICD BlockKeys.

    Built from two Excel traceability tables:
      - 设备->高层需求矩阵 (ERD -> HLR)
      - 设备->ICD追溯表 (ERD -> ICD FullName)
    """

    hlr_to_blocks: dict[str, set[str]] = field(default_factory=dict)
    hlr_to_erds: dict[str, list[str]] = field(default_factory=dict)
    erd_to_icd: dict[str, list[str]] = field(default_factory=dict)

    # Statistics
    total_hlrs_traced: int = 0
    total_erds: int = 0
    total_icd_fullnames: int = 0
    icd_mapped_to_blocks: int = 0
    icd_unmapped: list[str] = field(default_factory=list)


# -- Core mapping function -------------------------------------------

def name_to_block_key(name: str) -> str | None:
    """Convert an ICD FullName or signal_name to its ICDBlock.block_key.

    Mirrors the logic in signal_profiler._extract_profile_key +
    _extract_signal_family, but as a standalone function (zero imports
    from matching to avoid coupling).

    A429 example:
      "HF_AMSC2.pi429_B2_275_2.L275_2_B2_OVHD_CPA_PACK_FLOW_RECIRC_TRIM_AIR_R_PACK_2"
      -> "L275/OVHD_CPA_PACK_FLOW_RECIRC_TRIM_AIR_R_PACK_2"

    Non-label example (CAN/A825):
      "HF_AMSC2.pi825/SPEED_CMD"
      -> "pi825/SPEED_CMD"

    Returns None for empty or unparseable names.
    """
    if not name or not name.strip():
        return None

    segments = [s.strip() for s in name.split('.') if s.strip()]
    if not segments:
        return None

    label = None
    for seg in segments:
        m = _LABEL_SEGMENT_RE.match(seg)
        if m:
            label = m.group(1)
            break

    leaf = segments[-1]

    if label:
        label_value = label[1:]  # strip 'L'
        m2 = _SIGNAL_FAMILY_PREFIX_RE.match(leaf)
        if m2:
            family = leaf[m2.end():]
        else:
            family = leaf
        return f"L{label_value}/{family}"
    else:
        if len(segments) >= 2:
            parent = segments[-2]
            return f"{parent}/{leaf}"
        return leaf


# -- Table readers ---------------------------------------------------

def _read_table2_erd_to_hlr(trace_dir: Path) -> dict[str, list[str]]:
    """Read Table 2 (单模块需求矩阵分析) to build ERD -> HLR mapping.

    Table 2 expected filename: 单模块需求矩阵分析（设备2软件高层）-裁剪.xlsx
    Sheet: index 0 (Sheet1)
      Col A (0): ERD编号 (fill-forward for merged cells)
      Col D (3): 下级需求编号 (HLR ID or empty)
      Col E (4): 下级模块名称 (skip when "EICD")

    Returns: {erd_id: [hlr_id, ...]}
    """
    import openpyxl

    erd_to_hlr: dict[str, list[str]] = defaultdict(list)

    fname = "单模块需求矩阵分析（设备2软件高层）-裁剪.xlsx"
    fpath = trace_dir / fname
    if not fpath.exists():
        raise FileNotFoundError(f"Table 2 not found: {fpath}")

    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[0]]  # Sheet1

    current_erd = ""

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True), start=4):
        erd_cell = str(row[0]).strip() if row[0] else ""
        hlr_cell = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        module_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""

        # Fill-forward: non-empty ERD updates current
        if erd_cell and erd_cell != 'None':
            current_erd = erd_cell

        # Skip rows without a current ERD
        if not current_erd:
            continue

        # Skip rows whose module is EICD (direct ERD->ICD ref, not HLR)
        if module_name.upper() == "EICD":
            continue

        # Skip rows with no HLR ID
        if not hlr_cell or hlr_cell == 'None':
            continue

        erd_to_hlr[current_erd].append(hlr_cell)

    wb.close()
    return dict(erd_to_hlr)


def _read_table1_erd_to_icd(trace_dir: Path) -> dict[str, list[str]]:
    """Read Table 1 (设备需求与系统ICD追溯表) to build ERD -> ICD FullName mapping.

    Table 1 expected filename: 设备需求与系统ICD追溯表.xlsx
    Sheet: index 1 (设备_设备接口追溯表)
      Col D (3): ERD编号* (fill-forward for merged cells)
      Col H (7): ICD FullName*

    Returns: {erd_id: [icd_fullname, ...]}
    """
    import openpyxl

    erd_to_icd: dict[str, list[str]] = defaultdict(list)

    fname = "设备需求与系统ICD追溯表.xlsx"
    fpath = trace_dir / fname
    if not fpath.exists():
        raise FileNotFoundError(f"Table 1 not found: {fpath}")

    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[wb.sheetnames[1]]  # 设备_设备接口追溯表

    current_erd = ""

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        erd_cell = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        icd_fn = str(row[7]).strip() if len(row) > 7 and row[7] else ""

        if erd_cell and erd_cell != 'None':
            current_erd = erd_cell

        if not current_erd:
            continue

        if not icd_fn or icd_fn == 'None':
            continue

        erd_to_icd[current_erd].append(icd_fn)

    wb.close()
    return dict(erd_to_icd)


# -- Index builder ---------------------------------------------------

def build_trace_index(trace_dir: Path) -> TraceabilityIndex:
    """Build the complete HLR -> BlockKey traceability index.

    Args:
        trace_dir: Directory containing the two traceability Excel files.

    Returns:
        TraceabilityIndex with hlr_to_blocks, statistics, and raw mappings.

    Raises:
        FileNotFoundError: if either expected Excel file is missing.
    """
    # Step 1: Read both tables
    erd_to_hlr = _read_table2_erd_to_hlr(trace_dir)
    erd_to_icd = _read_table1_erd_to_icd(trace_dir)

    # Step 2: Combine -> HLR -> [ICD FullName, ...]
    index = TraceabilityIndex()
    index.hlr_to_erds = defaultdict(list)
    index.erd_to_icd = erd_to_icd
    index.total_erds = len(erd_to_hlr)  # ERDs that link to HLRs

    hlr_to_icd: dict[str, set[str]] = defaultdict(set)
    for erd, hlr_list in erd_to_hlr.items():
        icd_list = erd_to_icd.get(erd, [])
        for hlr in hlr_list:
            index.hlr_to_erds[hlr].append(erd)
            for icd_fn in icd_list:
                hlr_to_icd[hlr].add(icd_fn)

    index.total_hlrs_traced = len(hlr_to_icd)

    # Step 3: Map ICD FullName -> block_key
    all_icd_fullnames: set[str] = set()
    for icd_set in hlr_to_icd.values():
        all_icd_fullnames.update(icd_set)
    index.total_icd_fullnames = len(all_icd_fullnames)

    index.hlr_to_blocks = defaultdict(set)
    for icd_fn in all_icd_fullnames:
        bk = name_to_block_key(icd_fn)
        if bk:
            index.icd_mapped_to_blocks += 1
            # Assign this block_key to ALL HLRs that reference this ICD FullName
            for hlr, icd_set in hlr_to_icd.items():
                if icd_fn in icd_set:
                    index.hlr_to_blocks[hlr].add(bk)
        else:
            index.icd_unmapped.append(icd_fn)

    # Convert defaultdict to plain dict
    index.hlr_to_blocks = dict(index.hlr_to_blocks)
    index.hlr_to_erds = dict(index.hlr_to_erds)

    return index
